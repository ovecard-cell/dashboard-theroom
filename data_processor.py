import pandas as pd
import json
from pathlib import Path
from datetime import datetime, date

# Base dir: la carpeta donde está este archivo (funciona local y en Streamlit Cloud)
_BASE_DIR = Path(__file__).resolve().parent

# ── Clasificación de stock nuevo ──────────────────────────────────────────────

KAZUMA_DATE = pd.Timestamp("2026-03-26")
LISBON_DATE = pd.Timestamp("2026-03-28")
DISTRICT_DATE = pd.Timestamp("2026-03-30")

KAZUMA_PREFIXES = [
    "REMERA BASICA LYCRA TOM", "REMERA OVER POPPIES", "REMERA OVER LOS ANGELES",
    "REMERA OVER CORAL", "REMERA OVER EVERY", "REMERA REGULAR TACOS",
    "REMERA REGULAR GRAVE", "REMERA REGULAR JUICY", "CAMISA ROTHE",
    "JEAN RIG MOM MIROX", "JEAN RIG MOM JARO",
]
LISBON_PREFIXES = [
    "PANTALON RINGO", "JEAN SEVEN STRAIGHT", "JEAN FALCON REGULAR",
    "CHOMBA TEJIDA", "REMERA OVER LISA LB", "REMERA OVER TRIEND",
]
DISTRICT_PREFIXES = [
    "REMERA GIRONA", "REMERA NAPOLES", "JEAN CENTINELA",
    "CHOMBA EMINEM II", "BILLETERA",
]


def _classify_row(nombre: str, fecha: pd.Timestamp):
    """Devuelve (stock_tipo, proveedor_nuevo)."""
    nombre = str(nombre).upper().strip()
    if pd.notna(fecha):
        if fecha >= KAZUMA_DATE:
            for p in KAZUMA_PREFIXES:
                if nombre.startswith(p):
                    return "NUEVO", "KAZUMA"
        if fecha >= LISBON_DATE:
            for p in LISBON_PREFIXES:
                if nombre.startswith(p):
                    return "NUEVO", "LISBON"
        if fecha >= DISTRICT_DATE:
            for p in DISTRICT_PREFIXES:
                if nombre.startswith(p):
                    return "NUEVO", "DISTRICT"
    return "VIEJO", None


# ── Lectura de archivos Dux ───────────────────────────────────────────────────

def _leer_xls_directo(path) -> pd.DataFrame:
    """Lee un .xls de Dux usando xlrd directamente (sin pandas read_excel)."""
    import xlrd
    wb  = xlrd.open_workbook(str(path))
    ws  = wb.sheet_by_index(0)
    # Headers en fila 2 (índice 2)
    headers = [str(ws.cell_value(2, j)).strip() for j in range(ws.ncols)]
    rows = []
    for i in range(3, ws.nrows):
        row = {}
        for j, h in enumerate(headers):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row[h] = dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    row[h] = str(cell.value)
            else:
                row[h] = str(cell.value)
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def load_dux_files(folder: str = "data/ventas") -> pd.DataFrame:
    """
    Lee todos los .xls exportados por Dux en la carpeta indicada,
    aplica filtros y devuelve un DataFrame limpio.
    """
    path = _BASE_DIR / folder if not Path(folder).is_absolute() else Path(folder)
    archivos = list(path.glob("*.xls")) + list(path.glob("*.xlsx"))

    if not archivos:
        return pd.DataFrame()

    frames = []
    for f in archivos:
        try:
            if f.suffix == ".xls":
                df = _leer_xls_directo(f)
            else:
                df = pd.read_excel(f, header=2, engine="openpyxl", dtype=str)
            frames.append(df)
        except Exception as e:
            print(f"[WARN] No se pudo leer {f.name}: {e}")

    if not frames:
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)
    df.columns = df.columns.str.strip()

    # ── Normalizar columnas clave ─────────────────────────────────────────────
    col_map = {
        "Fecha Comp": "fecha_raw",
        "Producto": "producto",
        "Cantidad": "cantidad_raw",
        "Total Sin IVA": "neto_raw",
        "Total Con Iva": "total_con_iva_raw",
        "% Desc.": "descuento_raw",
        "Precio Uni": "precio_lista_raw",
        "Costo Unitario Producto": "costo_raw",
        "Forma Pago": "forma_pago",
        "Personal Registra": "personal",
        "Marca": "marca",
        "Rubro": "rubro",
        "Sub Rubro": "sub_rubro",
    }
    for orig, nuevo in col_map.items():
        if orig in df.columns:
            df[nuevo] = df[orig]
        elif nuevo not in df.columns:
            df[nuevo] = ""

    # ── Parsear tipos ─────────────────────────────────────────────────────────
    df["fecha"] = pd.to_datetime(df["fecha_raw"], dayfirst=True, errors="coerce")
    df["cantidad"] = pd.to_numeric(df["cantidad_raw"], errors="coerce").fillna(0)
    df["neto"] = pd.to_numeric(df["neto_raw"], errors="coerce").fillna(0)
    df["descuento"] = pd.to_numeric(df["descuento_raw"], errors="coerce").fillna(0)
    df["precio_lista"] = pd.to_numeric(df["precio_lista_raw"], errors="coerce").fillna(0)
    df["costo"] = pd.to_numeric(df["costo_raw"], errors="coerce").fillna(0)
    # Total con IVA: usar columna real si existe, sino calcular al 21%
    if "total_con_iva_raw" in df.columns:
        df["total_con_iva"] = pd.to_numeric(df["total_con_iva_raw"], errors="coerce").fillna(0)
        # Si la columna vino vacía, calcular
        if df["total_con_iva"].sum() == 0:
            df["total_con_iva"] = df["neto"] * 1.21
    else:
        df["total_con_iva"] = df["neto"] * 1.21

    # ── Filtros ───────────────────────────────────────────────────────────────
    # Quitar envíos
    df = df[~df["producto"].astype(str).str.upper().str.contains("ENVIO|COSTO_ENVIO", na=False)]
    # Quitar devoluciones y cero
    df = df[df["cantidad"] > 0]
    # Quitar sin fecha
    df = df[df["fecha"].notna()]
    # Deduplicar filas idénticas (si se subió el mismo archivo dos veces o con rangos que se pisan)
    df = df.drop_duplicates(subset=["fecha_raw", "producto", "cantidad_raw", "neto_raw", "forma_pago"])

    # ── Canal ─────────────────────────────────────────────────────────────────
    df["canal"] = df["personal"].astype(str).apply(
        lambda x: "Online" if "THEROOM1" in x.upper() else "Físico"
    )

    # ── Clasificación nuevo/viejo ─────────────────────────────────────────────
    clasificacion = df.apply(
        lambda r: _classify_row(str(r["producto"]), r["fecha"]),
        axis=1,
        result_type="expand",
    )
    df["stock_tipo"] = clasificacion[0]
    df["proveedor_nuevo"] = clasificacion[1]

    # ── Columnas de tiempo ────────────────────────────────────────────────────
    df["fecha_dia"] = df["fecha"].dt.date
    df["anio"] = df["fecha"].dt.year
    df["mes"] = df["fecha"].dt.month

    df = df.reset_index(drop=True)

    return df


# ── Helpers de análisis ───────────────────────────────────────────────────────

def ventas_por_dia(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["fecha_dia", "neto", "cantidad"])
    return (
        df.groupby("fecha_dia")
        .agg(neto=("neto", "sum"), cantidad=("cantidad", "sum"))
        .reset_index()
        .sort_values("fecha_dia")
    )


def ventas_hoy(df: pd.DataFrame) -> dict:
    hoy = date.today()
    sub = df[df["fecha_dia"] == hoy]
    return {
        "neto": sub["neto"].sum(),
        "cantidad": int(sub["cantidad"].sum()),
        "transacciones": len(sub),
    }


def ventas_mes(df: pd.DataFrame, anio: int = None, mes: int = None) -> dict:
    hoy = date.today()
    anio = anio or hoy.year
    mes = mes or hoy.month
    sub = df[(df["anio"] == anio) & (df["mes"] == mes)]
    dias_con_venta = sub["fecha_dia"].nunique()
    return {
        "neto": sub["neto"].sum(),
        "cantidad": int(sub["cantidad"].sum()),
        "dias_con_venta": dias_con_venta,
    }


def proyeccion_mes(df: pd.DataFrame) -> float:
    """Proyecta ventas del mes basado en el promedio diario hasta hoy."""
    hoy = date.today()
    datos_mes = ventas_mes(df)
    dias_con_venta = datos_mes["dias_con_venta"]
    if dias_con_venta == 0:
        return 0.0
    import calendar
    dias_totales = calendar.monthrange(hoy.year, hoy.month)[1]
    promedio = datos_mes["neto"] / dias_con_venta
    return promedio * dias_totales


def load_ventas_manuales(path: str = "data/ventas_manuales.json") -> pd.DataFrame:
    """Carga ventas ingresadas manualmente (sin Dux)."""
    p = _BASE_DIR / path if not Path(path).is_absolute() else Path(path)
    if not p.exists():
        return pd.DataFrame()
    data = __import__("json").loads(p.read_text(encoding="utf-8"))
    if not data:
        return pd.DataFrame()
    df = pd.DataFrame(data)
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["fecha_dia"] = df["fecha"].dt.date
    df["anio"] = df["fecha"].dt.year
    df["mes"] = df["fecha"].dt.month
    df["stock_tipo"] = "MANUAL"
    df["proveedor_nuevo"] = None
    df["canal"] = df.get("canal", "Físico")
    return df


def merge_ventas(df_dux: pd.DataFrame, df_manual: pd.DataFrame) -> pd.DataFrame:
    """
    Une ventas de Dux con ventas manuales.
    Si hay Dux, descarta días duplicados del manual (Dux tiene precedencia).
    """
    if df_dux.empty and df_manual.empty:
        return pd.DataFrame()
    if df_dux.empty:
        return df_manual
    if df_manual.empty:
        return df_dux
    # Fechas que ya tienen datos de Dux
    fechas_dux = set(df_dux["fecha_dia"].unique())
    df_manual_filtrado = df_manual[~df_manual["fecha_dia"].isin(fechas_dux)]
    return pd.concat([df_dux, df_manual_filtrado], ignore_index=True)


def load_compras_dux(path) -> list:
    """Lee el archivo de Consulta de Compras Detallada exportado de Dux."""
    import xlrd
    try:
        if isinstance(path, (str, Path)):
            wb = xlrd.open_workbook(str(path))
        else:
            wb = xlrd.open_workbook(file_contents=path)
    except Exception:
        return []
    ws = wb.sheet_by_index(0)

    # Headers en fila 2
    headers = []
    for j in range(ws.ncols):
        headers.append(str(ws.cell_value(2, j)).strip().upper())

    gastos = []
    for i in range(3, ws.nrows):
        try:
            tipo = str(ws.cell_value(i, 2)).strip().upper()
            if tipo not in ("GASTO", "COMPRA"):
                continue

            # Fecha
            cell_f = ws.cell(i, 7)
            if cell_f.ctype == xlrd.XL_CELL_DATE:
                dt = xlrd.xldate_as_datetime(cell_f.value, wb.datemode)
                fecha = dt.strftime("%Y-%m-%d")
            else:
                from datetime import datetime as _dt
                fecha_str = str(cell_f.value).strip()
                for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                    try:
                        fecha = _dt.strptime(fecha_str[:10], fmt).strftime("%Y-%m-%d")
                        break
                    except Exception:
                        fecha = ""

            proveedor = str(ws.cell_value(i, 3)).strip()
            producto = str(ws.cell_value(i, 11)).strip()
            monto = float(ws.cell_value(i, 21)) if ws.cell_value(i, 21) else 0
            neto = float(ws.cell_value(i, 16)) if ws.cell_value(i, 16) else monto
            iva = float(ws.cell_value(i, 18)) if ws.cell_value(i, 18) else 0
            efectivo = float(ws.cell_value(i, 23)) if ws.cell_value(i, 23) else 0
            cheque = float(ws.cell_value(i, 24)) if ws.cell_value(i, 24) else 0
            cuenta = float(ws.cell_value(i, 25)) if ws.cell_value(i, 25) else 0
            obs = str(ws.cell_value(i, 30)).strip() if ws.cell_value(i, 30) else ""

            if monto <= 0:
                continue

            # Determinar medio de pago
            if efectivo > 0:
                medio = "Efectivo"
            elif cuenta > 0:
                medio = "Cuenta bancaria"
            elif cheque > 0:
                medio = "Cheque"
            else:
                medio = "Sin especificar"

            # Categorizar
            prod_up = producto.upper()
            prov_up = proveedor.upper()
            if tipo == "COMPRA":
                categoria = "Mercadería"
            elif "LUZ" in prod_up or "GAS" in prod_up or "TEL" in prod_up or "DPEC" in prov_up or "AGUA" in prov_up or "ENERGIA" in prov_up:
                categoria = "Servicios (luz/gas/tel)"
            elif "GABRIELA" in prod_up or "SOFIA" in prod_up or "SUELDO" in prod_up or "GABRIELA" in prov_up:
                categoria = "Sueldos"
            elif "TRANSPORTE" in prod_up or "ENVIO" in prod_up or "CORREO" in prov_up:
                categoria = "Transporte"
            elif "ALQUILER" in prod_up:
                categoria = "Alquiler"
            elif "IMPUESTO" in prod_up or "AFIP" in prod_up or "ARCA" in prod_up:
                categoria = "AFIP / Impuestos"
            else:
                categoria = "Otros"

            concepto = f"{proveedor} — {producto}"
            if obs:
                concepto += f" ({obs[:40]})"

            gastos.append({
                "fecha": fecha,
                "concepto": concepto,
                "monto": monto,
                "neto": neto,
                "iva": iva,
                "categoria": categoria,
                "medio": medio,
                "notas": f"Importado de Dux compras",
            })
        except Exception:
            pass

    return gastos


def load_stock_dux(path: str = "data/stock_dux.xls") -> pd.DataFrame:
    """Lee el archivo de Valorización de Stock exportado de Dux."""
    p = _BASE_DIR / path if not Path(path).is_absolute() else Path(path)
    if not p.exists():
        return pd.DataFrame()
    import xlrd
    wb = xlrd.open_workbook(str(p))
    ws = wb.sheet_by_index(0)
    productos = []
    for i in range(3, ws.nrows):
        try:
            prod = str(ws.cell_value(i, 1)).strip()
            rubro = str(ws.cell_value(i, 3)).strip()
            sub = str(ws.cell_value(i, 4)).strip()
            cant = float(ws.cell_value(i, 5)) if ws.cell_value(i, 5) else 0
            prov = str(ws.cell_value(i, 7)).strip()
            val_unit = float(ws.cell_value(i, 8)) if ws.cell_value(i, 8) else 0
            val_total = float(ws.cell_value(i, 9)) if ws.cell_value(i, 9) else 0
            if prod:
                productos.append({
                    "producto": prod, "rubro": rubro, "sub_rubro": sub,
                    "cantidad": int(cant), "proveedor_dux": prov,
                    "costo_unit": val_unit, "valor_total": val_total,
                })
        except Exception:
            pass
    return pd.DataFrame(productos)


def stock_nuevo_resumen(df: pd.DataFrame, stock_inicial: list) -> pd.DataFrame:
    """
    Cruza stock_inicial con ventas para calcular velocidad y días restantes.
    Si no hay .xls, usa el campo vendidos_manual del stock_inicial.
    """
    df_valido = not df.empty and "stock_tipo" in df.columns and "producto" in df.columns

    filas = []
    for item in stock_inicial:
        proveedor = item["proveedor"]
        tipo = item["tipo"]
        stock_ini = item["stock_inicial"]
        fecha_ing = pd.Timestamp(item["fecha_ingreso"])
        prefijos = [p.upper() for p in item["prefijos"]]
        vendidos_manual_fallback = item.get("vendidos_manual", 0)

        if df_valido:
            mask_nuevo = df["stock_tipo"] == "NUEVO"
            if mask_nuevo.any():
                prod_upper = df["producto"].astype(str).str.upper().fillna("")
                mask_prod = prod_upper.str.startswith(prefijos[0]) if prefijos else pd.Series(False, index=df.index)
                for p in prefijos[1:]:
                    mask_prod = mask_prod | prod_upper.str.startswith(p)
                mask = mask_nuevo & (df["proveedor_nuevo"] == proveedor) & mask_prod
                vendidos = int(df.loc[mask, "cantidad"].sum())
                neto_generado = df.loc[mask, "neto"].sum()
            else:
                vendidos = vendidos_manual_fallback
                neto_generado = 0.0
        else:
            vendidos = vendidos_manual_fallback
            neto_generado = 0.0

        dias_transcurridos = max((date.today() - fecha_ing.date()).days, 1)
        vel_dia = vendidos / dias_transcurridos
        quedan = stock_ini - vendidos
        dias_stock = round(quedan / vel_dia) if vel_dia > 0 else 9999

        filas.append({
            "Proveedor": proveedor,
            "Tipo": tipo,
            "Stock Inicial": stock_ini,
            "Vendidos": vendidos,
            "Quedan": quedan,
            "Neto $": neto_generado,
            "Vel/día": round(vel_dia, 2),
            "Días de stock": dias_stock if dias_stock < 9999 else "∞",
            "% Vendido": round(vendidos / stock_ini * 100, 1) if stock_ini > 0 else 0,
            "Días transcurridos": dias_transcurridos,
        })

    return pd.DataFrame(filas)


# ── Parser extracto Banco Corrientes ─────────────────────────────────────────

def _parsear_importe(valor) -> float:
    """Convierte un valor a float: soporta floats de Python, '1.234.567,89' argentino, etc."""
    # Si ya es número, devolver directo
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        s = str(valor).strip().replace("$", "").replace(" ", "")
        if not s or s in ("", "-", "None", "none"):
            return 0.0
        # Si tiene coma → formato argentino (1.234.567,89)
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
            return float(s)
        # Si tiene un solo punto → float estándar (1520000.0)
        # Si tiene múltiples puntos → separador de miles sin decimales (1.234.567)
        puntos = s.count(".")
        if puntos <= 1:
            return float(s)
        else:
            s = s.replace(".", "")
            return float(s)
    except Exception:
        return 0.0


def _categorizar_movimiento(tipo: str, referencia: str) -> tuple:
    """Devuelve (categoria, descripcion) para cada movimiento."""
    tipo_u = str(tipo).upper().strip()
    ref_u  = str(referencia).upper().strip()

    # Impuestos
    if "25413" in tipo_u or "25.4" in tipo_u:
        return "AFIP / Impuestos", "Impuesto al cheque (Ley 25413)"
    elif "IMPUESTO LEY" in tipo_u:
        return "AFIP / Impuestos", f"Impuesto Ley — {tipo.strip()}"
    elif "IVA TASA" in tipo_u:
        return "AFIP / Impuestos", f"IVA — {tipo.strip()}"
    elif "PERCEPCION I" in tipo_u or "PER. CORRIEN" in tipo_u:
        return "AFIP / Impuestos", f"Percepcion IIBB — {tipo.strip()}"
    elif "SELLADO" in tipo_u:
        return "AFIP / Impuestos", f"Sellado — {tipo.strip()}"
    elif "ARCA" in tipo_u:
        return "AFIP / Impuestos", f"ARCA (AFIP) — {tipo.strip()}"
    # Transferencias
    elif "TRANSF" in tipo_u or "TRF" in tipo_u:
        return "TRANSFERENCIA", f"Transferencia — {tipo.strip()}"
    # Tarjeta de crédito
    elif "T.CREDITO" in tipo_u or "T.DEBITO" in tipo_u or "COBRO T." in tipo_u:
        return "Gastos bancarios", f"Pago tarjeta — {tipo.strip()}"
    # Cheques
    elif "CHEQUE" in tipo_u or "CHQ" in tipo_u:
        return "Cheque debitado", f"Cheque — {ref_u or tipo.strip()}"
    # Comisiones y mantenimiento bancario
    elif "COM MANT" in tipo_u or "COMISION" in tipo_u or "MANTENIMIENTO" in tipo_u:
        return "Gastos bancarios", f"Comision bancaria — {tipo.strip()}"
    elif "INTERES COBR" in tipo_u or "AJUSTE DE IN" in tipo_u:
        return "Gastos bancarios", f"Intereses bancarios — {tipo.strip()}"
    # Débitos automáticos (alarmas, servicios)
    elif "OG-DEBITO" in tipo_u or "DEBITO DI" in tipo_u:
        return "Gastos bancarios", f"Débito automático — {tipo.strip()}"
    # Acreditaciones
    elif "ACREDITACION" in tipo_u:
        return "CREDITO", f"Acreditacion — {tipo.strip()}"
    else:
        return "Gastos bancarios", tipo.strip() if tipo.strip() else "Movimiento bancario"


def parsear_extracto_corrientes(file_bytes: bytes, nombre_archivo: str) -> dict:
    """
    Lee un .xls exportado por el Banco Corrientes.
    Devuelve dict con:
      - saldo_actual: float
      - movimientos: list[dict]  (fecha, tipo, importe, debcred, categoria, descripcion)
      - gastos_nuevos: list[dict] para agregar a gastos.json
      - resumen: dict con totales
    """
    import xlrd, io

    # Intentar primero con xlrd (.xls), luego openpyxl (.xlsx), luego HTML (bancos que exportan HTML como .xls)
    wb = None
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
    except Exception:
        pass

    if wb is None:
        try:
            from openpyxl import load_workbook
            wb_xl = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws_xl = wb_xl.active
            rows_raw = [[str(c.value or "") for c in row] for row in ws_xl.iter_rows()]
            return _parsear_extracto_rows(rows_raw, nombre_archivo)
        except Exception:
            pass

    if wb is None:
        # Muchos bancos argentinos (Santander, Galicia, MP) exportan archivos .xls
        # que en realidad son HTML o CSV. Probar ambos formatos.
        import pandas as pd

        # Intentar como CSV primero (Mercado Pago exporta CSV con extensión .xls)
        for encoding in ("utf-8", "latin-1"):
            try:
                contenido = file_bytes.decode(encoding, errors="replace")
                if "," in contenido[:500] or ";" in contenido[:500]:
                    sep = ";" if contenido.count(";") > contenido.count(",") else ","
                    df_csv = pd.read_csv(io.StringIO(contenido), sep=sep)
                    if len(df_csv.columns) >= 3 and len(df_csv) >= 1:
                        rows_raw = []
                        rows_raw.append([str(c) for c in df_csv.columns.tolist()])
                        for _, fila in df_csv.iterrows():
                            rows_raw.append([str(v) if str(v) != "nan" else "" for v in fila.tolist()])
                        return _parsear_extracto_rows(rows_raw, nombre_archivo)
            except Exception:
                pass

        # Intentar como HTML (Santander, Galicia exportan HTML disfrazado de .xls)
        for encoding in ("utf-8", "latin-1"):
            try:
                contenido = file_bytes.decode(encoding, errors="replace")
                tablas = pd.read_html(io.StringIO(contenido))
                if not tablas:
                    continue
                rows_raw = []
                for tbl in tablas:
                    rows_raw.append([str(c) for c in tbl.columns.tolist()])
                    for _, fila in tbl.iterrows():
                        rows_raw.append([str(v) if str(v) != "nan" else "" for v in fila.tolist()])
                return _parsear_extracto_rows(rows_raw, nombre_archivo)
            except Exception:
                pass

        return {"error": "No se pudo leer el archivo. Probá exportarlo de nuevo desde el banco como Excel o CSV."}

    ws = wb.sheet_by_index(0)
    rows_raw = []
    for i in range(ws.nrows):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(dt.strftime("%d/%m/%Y"))
                except Exception:
                    row.append(str(cell.value))
            else:
                row.append(str(cell.value).strip())
        rows_raw.append(row)

    return _parsear_extracto_rows(rows_raw, nombre_archivo)


def _parsear_extracto_rows(rows_raw: list, nombre_archivo: str) -> dict:
    """Procesa las filas crudas del extracto y devuelve el resultado estructurado."""
    from datetime import datetime as _dt

    saldo_actual = 0.0
    movimientos  = []
    gastos_nuevos = []

    # Buscar saldo actual y la tabla de movimientos
    header_idx = None
    for i, row in enumerate(rows_raw):
        fila = " ".join(row).upper()

        # Buscar saldo: "SALDO ACTUAL", "SALDO:" etc.
        if ("SALDO ACTUAL" in fila or (fila.startswith("SALDO") and ":" in fila)) and saldo_actual == 0.0:
            for cell in row:
                v = _parsear_importe(cell)
                if abs(v) > 100:
                    saldo_actual = v
                    break

        # Buscar fila de headers
        if "FECHA" in fila and (
            "TIPO" in fila or "TRANSACCION" in fila or "IMPORTE" in fila
            or "CONCEPTO" in fila or "DETALLE" in fila
            or "CREDITO" in fila or "DEBITO" in fila
        ):
            header_idx = i
            break

    if header_idx is None:
        return {"error": "No se encontro la tabla de movimientos en el archivo"}

    # Detectar columnas por encabezado (quitar acentos y chars raros)
    import unicodedata, re
    def _limpiar(s):
        s = str(s).upper().strip()
        # Normalizar unicode y quitar acentos
        nfkd = unicodedata.normalize('NFKD', s)
        s = "".join(c for c in nfkd if not unicodedata.combining(c))
        # Quitar chars no-ASCII restantes
        s = re.sub(r'[^\x20-\x7E]', '', s)
        return s

    headers = [_limpiar(h) for h in rows_raw[header_idx]]

    def _col(keys):
        for k in keys:
            for j, h in enumerate(headers):
                if k in h:
                    return j
        return None

    col_fecha  = _col(["FECHA"])
    col_tipo   = _col(["CONCEPTO", "TIPO", "TRANSACCION", "DESCRIPCION", "DETALLE"])
    col_ref    = _col(["REFERENCIA", "REF", "NUMERO DOC", "DOCUMENTO"])
    col_dc     = _col(["D/C", "DC"])
    col_imp    = _col(["IMPORTE", "MONTO", "VALOR"])
    col_cred   = _col(["CREDITO", "CRED"])
    col_deb    = _col(["DEBITO", "DEB"])

    # Formato BBVA: columnas separadas Crédito/Débito (no hay col "Importe" unica)
    tiene_cols_separadas = col_cred is not None or col_deb is not None

    if col_fecha is None:
        return {"error": "No se encontro la columna Fecha en el extracto"}

    if col_imp is None and not tiene_cols_separadas:
        return {"error": "No se encontraron columnas de importe en el extracto"}

    for row in rows_raw[header_idx + 1:]:
        if not row or len(row) <= col_fecha:
            continue

        fecha_str = str(row[col_fecha]).strip()
        if not fecha_str or not any(c.isdigit() for c in fecha_str):
            continue

        tipo_str = str(row[col_tipo]).strip() if col_tipo is not None and col_tipo < len(row) else ""
        ref_str  = str(row[col_ref]).strip()  if col_ref  is not None and col_ref  < len(row) else ""

        # Determinar importe y si es débito
        if tiene_cols_separadas:
            # BBVA: crédito y débito en columnas separadas
            cred_val = _parsear_importe(row[col_cred]) if col_cred is not None and col_cred < len(row) else 0.0
            deb_val  = _parsear_importe(row[col_deb])  if col_deb  is not None and col_deb  < len(row) else 0.0
            if abs(deb_val) > 0:
                importe   = abs(deb_val)
                es_debito = True
            elif abs(cred_val) > 0:
                importe   = abs(cred_val)
                es_debito = False
            else:
                continue
        else:
            # Corrientes: importe unico + columna D/C
            imp_str = str(row[col_imp]) if col_imp < len(row) else "0"
            importe = abs(_parsear_importe(imp_str))
            if importe == 0:
                continue
            if col_dc is not None and col_dc < len(row):
                dc_str = str(row[col_dc]).upper().strip()
                es_debito = "DEB" in dc_str or dc_str == "D"
            else:
                es_debito = True  # por defecto

        # Parsear fecha
        fecha_parsed = None
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                fecha_parsed = _dt.strptime(fecha_str[:10], fmt).date()
                break
            except Exception:
                pass
        if fecha_parsed is None:
            continue

        categoria, descripcion = _categorizar_movimiento(tipo_str, ref_str)

        mov = {
            "fecha":       fecha_parsed.isoformat(),
            "tipo":        tipo_str,
            "referencia":  ref_str,
            "importe":     importe,
            "es_debito":   es_debito,
            "categoria":   categoria,
            "descripcion": descripcion,
        }
        movimientos.append(mov)

        # Solo los débitos que son gastos reales van a gastos.json
        if es_debito and categoria not in ("TRANSFERENCIA", "CREDITO", "Cheque debitado"):
            gastos_nuevos.append({
                "fecha":      fecha_parsed.isoformat(),
                "concepto":   descripcion,
                "monto":      importe,
                "categoria":  categoria,
                "medio":      nombre_archivo,
                "notas":      f"Importado de {nombre_archivo}",
            })

    total_debitos  = sum(m["importe"] for m in movimientos if m["es_debito"])
    total_creditos = sum(m["importe"] for m in movimientos if not m["es_debito"])

    return {
        "saldo_actual":  saldo_actual,
        "movimientos":   movimientos,
        "gastos_nuevos": gastos_nuevos,
        "resumen": {
            "total_debitos":  total_debitos,
            "total_creditos": total_creditos,
            "cant_movimientos": len(movimientos),
        },
    }
