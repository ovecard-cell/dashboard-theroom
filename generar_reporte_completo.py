"""
Generador de reporte ejecutivo completo de The Room Argentina.
Arma un Excel con hojas: Resumen, Ventas año, Cheques, Bancos, Deuda personal,
Stock, Compras, Gastos.
"""
import json
import io
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


# Estilos globales
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1a1a2e")
SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="3a86ff")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="444455")
TOTAL_FILL = PatternFill("solid", fgColor="D0D0D0")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _style_title(ws, cell_range, text):
    first = cell_range.split(":")[0]
    ws.merge_cells(cell_range)
    ws[first] = text
    ws[first].font = TITLE_FONT
    ws[first].fill = TITLE_FILL
    ws[first].alignment = CENTER


def _style_section(ws, cell_range, text):
    first = cell_range.split(":")[0]
    ws.merge_cells(cell_range)
    ws[first] = text
    ws[first].font = SECTION_FONT
    ws[first].fill = SECTION_FILL
    ws[first].alignment = Alignment(horizontal="left", vertical="center")


def generar_reporte(hoy=None):
    hoy = hoy or date.today()
    wb = Workbook()

    # ======================================================================
    # Cargar datos
    # ======================================================================
    def _load(path, default):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return default

    bancos = _load("data/bancos.json", [])
    gastos = _load("data/gastos.json", [])
    cheques = _load("data/cheques.json", [])
    deuda_personal = _load("data/deuda_personal.json", {})
    stock_inicial = _load("data/stock_inicial.json", [])
    compras = _load("data/compras_mercaderia.json", [])

    # Cargar ventas Dux (vía data_processor)
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    try:
        from data_processor import load_dux_files, load_stock_dux
        df_ventas = load_dux_files()
        df_stock = load_stock_dux("data/stock_dux.xls")
    except Exception:
        import pandas as pd
        df_ventas = pd.DataFrame()
        df_stock = pd.DataFrame()

    # ======================================================================
    # HOJA 1: RESUMEN EJECUTIVO
    # ======================================================================
    ws = wb.active
    ws.title = "Resumen"
    _style_title(ws, "A1:G1", f"THE ROOM ARGENTINA — Resumen ejecutivo — {hoy.strftime('%d/%m/%Y')}")
    ws.row_dimensions[1].height = 30

    r = 3
    _style_section(ws, f"A{r}:G{r}", "SITUACIÓN HOY")
    r += 2

    # Situación: caja, ventas mes, stock, deuda
    ult_banco = bancos[-1] if bancos else {}
    caja = ult_banco.get("Efectivo Caja", 0)
    sant = ult_banco.get("Santander", 0)
    mp = ult_banco.get("Mercado Pago", 0)
    galicia = ult_banco.get("Galicia", 0)
    bbva = ult_banco.get("BBVA Frances", 0)
    ctes = ult_banco.get("Banco Corrientes", 0)
    disponible = max(sant, 0) + max(mp, 0) + max(galicia, 0) + caja
    descubierto = abs(min(bbva, 0)) + abs(min(ctes, 0))

    # Ventas del mes
    if not df_ventas.empty:
        df_mes = df_ventas[(df_ventas["fecha"].dt.month == hoy.month) & (df_ventas["fecha"].dt.year == hoy.year)]
        venta_mes_neto = float(df_mes["neto"].sum())
        venta_mes_iva = float(df_mes["total_con_iva"].sum())
        uds_mes = int(df_mes["cantidad"].sum())
        dias_mes = (hoy.day)
        prom_dia = venta_mes_iva / dias_mes if dias_mes else 0
    else:
        venta_mes_neto = venta_mes_iva = 0
        uds_mes = 0
        prom_dia = 0

    # Stock
    stock_uds = int(df_stock["cantidad"].sum()) if not df_stock.empty else 0
    stock_valor = float(df_stock["valor_total"].sum()) if not df_stock.empty else 0

    # Gastos mes
    gastos_mes = sum(
        g["monto"] for g in gastos
        if g["fecha"][:7] == f"{hoy.year}-{hoy.month:02d}"
    )
    cheques_pend = sum(c["monto"] for c in cheques if c.get("tipo") == "emitido" and c.get("estado") == "pendiente")

    situacion = [
        ("Disponible real", f"${disponible:,.0f}", f"Santander ${sant:,.0f} + MP ${mp:,.0f} + Caja ${caja:,.0f} + Galicia ${galicia:,.0f}"),
        ("Descubierto", f"-${descubierto:,.0f}", f"BBVA ${bbva:,.0f} + Corrientes ${ctes:,.0f}"),
        ("Posición neta", f"${disponible - descubierto:,.0f}", "Disponible menos descubierto"),
        ("", "", ""),
        ("Ventas del mes con IVA", f"${venta_mes_iva:,.0f}", f"{uds_mes} uds · {hoy.day} días · promedio ${prom_dia:,.0f}/día"),
        ("Ventas del mes sin IVA", f"${venta_mes_neto:,.0f}", "Neto facturado"),
        ("Gastos del mes", f"${gastos_mes:,.0f}", f"{len([g for g in gastos if g['fecha'][:7] == f'{hoy.year}-{hoy.month:02d}'])} movimientos"),
        ("", "", ""),
        ("Stock actual", f"{stock_uds} uds · ${stock_valor:,.0f}", "Valor de inventario según Dux"),
        ("Cheques pendientes", f"${cheques_pend:,.0f}", f"{len([c for c in cheques if c.get('tipo')=='emitido' and c.get('estado')=='pendiente'])} cheques emitidos sin pagar"),
    ]

    ws.cell(row=r, column=1, value="CONCEPTO").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="VALOR").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    ws.cell(row=r, column=3, value="DETALLE").font = HEADER_FONT
    ws.cell(row=r, column=3).fill = HEADER_FILL
    for c in (1, 2, 3):
        ws.cell(row=r, column=c).alignment = CENTER
    r += 1
    for concepto, valor, detalle in situacion:
        if concepto:
            ws.cell(row=r, column=1, value=concepto).font = Font(bold=True)
            ws.cell(row=r, column=2, value=valor).alignment = RIGHT
            ws.cell(row=r, column=3, value=detalle).font = Font(color="666666", italic=True)
        r += 1
    r += 1

    # Cheques urgentes
    _style_section(ws, f"A{r}:G{r}", "CHEQUES PRÓXIMOS A VENCER (30 días)")
    r += 2
    ws.cell(row=r, column=1, value="ID").font = HEADER_FONT
    ws.cell(row=r, column=2, value="PROVEEDOR").font = HEADER_FONT
    ws.cell(row=r, column=3, value="VENCE").font = HEADER_FONT
    ws.cell(row=r, column=4, value="MONTO").font = HEADER_FONT
    ws.cell(row=r, column=5, value="ESTADO").font = HEADER_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = HEADER_FILL
        ws.cell(row=r, column=c).alignment = CENTER
    r += 1

    pendientes_ord = []
    for c in cheques:
        if c.get("tipo") == "emitido" and c.get("estado") == "pendiente":
            try:
                venc = date.fromisoformat(c["vencimiento"])
                dias = (venc - hoy).days
                if dias <= 30:
                    pendientes_ord.append((dias, c))
            except Exception:
                pass
    pendientes_ord.sort(key=lambda x: x[0])
    total_prox = 0
    for dias, c in pendientes_ord:
        ws.cell(row=r, column=1, value=c["id"]).font = Font(bold=True)
        ws.cell(row=r, column=2, value=c["proveedor"])
        ws.cell(row=r, column=3, value=c["vencimiento"])
        ws.cell(row=r, column=4, value=float(c["monto"])).number_format = '"$"#,##0'
        if dias < 0:
            estado = f"VENCIDO hace {-dias}d"
            color_est = "FFCCCC"
        elif dias <= 7:
            estado = f"Vence en {dias}d"
            color_est = "FFE0CC"
        else:
            estado = f"En {dias}d"
            color_est = "FFF2CC"
        est_cell = ws.cell(row=r, column=5, value=estado)
        est_cell.fill = PatternFill("solid", fgColor=color_est)
        est_cell.alignment = CENTER
        total_prox += c["monto"]
        r += 1

    # Totales
    ws.cell(row=r, column=1, value="TOTAL 30 DÍAS").font = Font(bold=True)
    ws.cell(row=r, column=4, value=float(total_prox)).number_format = '"$"#,##0'
    ws.cell(row=r, column=4).font = Font(bold=True)
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    r += 2

    # Plata que el negocio debe al dueño
    if deuda_personal:
        _style_section(ws, f"A{r}:G{r}", "PLATA QUE EL NEGOCIO TE DEBE")
        r += 2
        if isinstance(deuda_personal, dict):
            for k, v in deuda_personal.items():
                if isinstance(v, (int, float)):
                    ws.cell(row=r, column=1, value=str(k).replace("_", " ").title()).font = Font(bold=True)
                    ws.cell(row=r, column=2, value=float(v)).number_format = '"$"#,##0'
                    r += 1
                elif isinstance(v, list) and v:
                    ws.cell(row=r, column=1, value=str(k).replace("_", " ").title()).font = Font(bold=True, color="3a86ff")
                    r += 1
                    for item in v:
                        if isinstance(item, dict):
                            concepto = item.get("concepto", "-")
                            fecha = item.get("fecha", "")
                            monto = item.get("monto", 0)
                            ws.cell(row=r, column=1, value=f"  {concepto}")
                            ws.cell(row=r, column=2, value=fecha)
                            ws.cell(row=r, column=3, value=float(monto)).number_format = '"$"#,##0'
                            r += 1
        r += 1

    # Ventas por año (si tenemos df)
    if not df_ventas.empty:
        _style_section(ws, f"A{r}:G{r}", f"VENTAS {hoy.year} — POR MES")
        r += 2
        ws.cell(row=r, column=1, value="MES").font = HEADER_FONT
        ws.cell(row=r, column=2, value="UDS").font = HEADER_FONT
        ws.cell(row=r, column=3, value="NETO").font = HEADER_FONT
        ws.cell(row=r, column=4, value="CON IVA").font = HEADER_FONT
        ws.cell(row=r, column=5, value="PROM/DÍA").font = HEADER_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = HEADER_FILL
            ws.cell(row=r, column=c).alignment = CENTER
        r += 1
        df_anio = df_ventas[df_ventas["fecha"].dt.year == hoy.year]
        for m in range(1, 13):
            df_m = df_anio[df_anio["fecha"].dt.month == m]
            if df_m.empty:
                continue
            uds = int(df_m["cantidad"].sum())
            neto = float(df_m["neto"].sum())
            iva = float(df_m["total_con_iva"].sum())
            dias_m = df_m["fecha"].dt.date.nunique()
            prom = iva / dias_m if dias_m else 0
            nombres_mes = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            ws.cell(row=r, column=1, value=nombres_mes[m-1]).font = Font(bold=True)
            ws.cell(row=r, column=2, value=uds).alignment = CENTER
            ws.cell(row=r, column=3, value=neto).number_format = '"$"#,##0'
            ws.cell(row=r, column=4, value=iva).number_format = '"$"#,##0'
            ws.cell(row=r, column=5, value=prom).number_format = '"$"#,##0'
            r += 1

    # Anchos
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A3"

    # ======================================================================
    # HOJA 2: VENTAS AÑO ACTUAL
    # ======================================================================
    ws2 = wb.create_sheet(f"Ventas {hoy.year}")
    _style_title(ws2, "A1:G1", f"Ventas {hoy.year} — actualizado al {hoy.strftime('%d/%m/%Y')}")
    ws2.row_dimensions[1].height = 28

    headers = ["Fecha", "Producto", "Cantidad", "Neto", "Con IVA", "Canal", "Forma pago"]
    ws2.append([])
    ws2.append(headers)
    _style_header_row(ws2, 3, 7)

    if not df_ventas.empty:
        df_a = df_ventas[df_ventas["fecha"].dt.year == hoy.year].sort_values("fecha", ascending=False)
        for _, r_ in df_a.iterrows():
            ws2.append([
                str(r_["fecha"])[:10],
                r_["producto"],
                int(r_["cantidad"]),
                float(r_["neto"]),
                float(r_["total_con_iva"]),
                r_["canal"],
                str(r_.get("forma_pago", ""))[:40],
            ])
        # Totales
        ws2.append(["TOTAL", "", int(df_a["cantidad"].sum()), float(df_a["neto"].sum()), float(df_a["total_con_iva"].sum()), "", ""])
        last = ws2.max_row
        for c in range(1, 8):
            ws2.cell(row=last, column=c).font = Font(bold=True)
            ws2.cell(row=last, column=c).fill = TOTAL_FILL

    for row_c in ws2.iter_rows(min_row=4, max_col=7):
        for cell in row_c:
            if cell.column in (4, 5):
                cell.number_format = '"$"#,##0'
    for col, w in zip(["A","B","C","D","E","F","G"], [12, 55, 10, 14, 14, 12, 40]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A4"

    # ======================================================================
    # HOJA 3: CHEQUES
    # ======================================================================
    ws3 = wb.create_sheet("Cheques")
    _style_title(ws3, "A1:G1", f"Cheques — actualizado al {hoy.strftime('%d/%m/%Y')}")
    ws3.row_dimensions[1].height = 28
    ws3.append([])
    ws3.append(["ID", "Proveedor", "Concepto", "Vencimiento", "Monto", "Tipo", "Estado"])
    _style_header_row(ws3, 3, 7)

    for c in sorted(cheques, key=lambda x: (x.get("estado","pendiente"), x.get("vencimiento","9999"))):
        venc = c.get("vencimiento", "")
        try:
            venc_d = date.fromisoformat(venc) if venc else None
            dias = (venc_d - hoy).days if venc_d else None
        except Exception:
            dias = None

        estado_txt = c.get("estado", "")
        if estado_txt == "pagado":
            estado_txt = f"PAGADO ({c.get('fecha_pago_real','')})"
            fill = PatternFill("solid", fgColor="CCFFCC")
        elif c.get("tipo") == "por_negociar":
            estado_txt = "POR NEGOCIAR"
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif dias is not None and dias < 0:
            estado_txt = f"VENCIDO hace {-dias}d"
            fill = PatternFill("solid", fgColor="FFCCCC")
        elif dias is not None and dias <= 7:
            estado_txt = f"Vence en {dias}d"
            fill = PatternFill("solid", fgColor="FFE0CC")
        else:
            estado_txt = f"Pendiente ({dias}d)" if dias is not None else "Pendiente"
            fill = PatternFill("solid", fgColor="E0E0F0")

        ws3.append([
            c.get("id", "-"),
            c.get("proveedor", ""),
            c.get("concepto", ""),
            venc,
            float(c.get("monto", 0)),
            c.get("tipo", ""),
            estado_txt,
        ])
        last = ws3.max_row
        ws3.cell(row=last, column=7).fill = fill

    # Totales por estado
    ws3.append([])
    emitidos_p = sum(c["monto"] for c in cheques if c.get("tipo")=="emitido" and c.get("estado")=="pendiente")
    emitidos_pag = sum(c["monto"] for c in cheques if c.get("tipo")=="emitido" and c.get("estado")=="pagado")
    negoc = sum(c["monto"] for c in cheques if c.get("tipo")=="por_negociar")
    ws3.append(["", "", "", "TOTAL EMITIDOS PENDIENTES", emitidos_p, "", ""])
    ws3.append(["", "", "", "TOTAL EMITIDOS PAGADOS", emitidos_pag, "", ""])
    ws3.append(["", "", "", "TOTAL POR NEGOCIAR", negoc, "", ""])
    for r_ in range(ws3.max_row - 2, ws3.max_row + 1):
        for c in range(1, 8):
            ws3.cell(row=r_, column=c).font = Font(bold=True)
            ws3.cell(row=r_, column=c).fill = TOTAL_FILL

    for row_c in ws3.iter_rows(min_row=4, max_col=7):
        for cell in row_c:
            if cell.column == 5:
                cell.number_format = '"$"#,##0'
    for col, w in zip(["A","B","C","D","E","F","G"], [8, 22, 40, 14, 14, 14, 20]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A4"

    # ======================================================================
    # HOJA 4: BANCOS
    # ======================================================================
    ws4 = wb.create_sheet("Bancos")
    _style_title(ws4, "A1:G1", f"Bancos — histórico de saldos")
    ws4.row_dimensions[1].height = 28
    ws4.append([])
    ws4.append(["Fecha", "BBVA Frances", "Corrientes", "Galicia", "Santander", "MP", "Caja"])
    _style_header_row(ws4, 3, 7)

    for b in bancos:
        ws4.append([
            b.get("fecha", ""),
            float(b.get("BBVA Frances", 0)),
            float(b.get("Banco Corrientes", 0)),
            float(b.get("Galicia", 0)),
            float(b.get("Santander", 0)),
            float(b.get("Mercado Pago", 0)),
            float(b.get("Efectivo Caja", 0)),
        ])
    for row_c in ws4.iter_rows(min_row=4, max_col=7):
        for cell in row_c:
            if cell.column >= 2:
                cell.number_format = '"$"#,##0'
                if cell.value is not None and cell.value < 0:
                    cell.font = Font(color="CC0000")
    for col, w in zip(["A","B","C","D","E","F","G"], [14, 16, 16, 14, 14, 14, 14]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A4"

    # ======================================================================
    # HOJA 5: STOCK ACTUAL
    # ======================================================================
    ws5 = wb.create_sheet("Stock")
    _style_title(ws5, "A1:E1", f"Stock actual — al {hoy.strftime('%d/%m/%Y')}")
    ws5.row_dimensions[1].height = 28
    ws5.append([])
    ws5.append(["Producto", "Proveedor", "Cantidad", "Costo unitario", "Valor total"])
    _style_header_row(ws5, 3, 5)

    if not df_stock.empty:
        df_stk_ord = df_stock.sort_values(["proveedor_dux", "cantidad"], ascending=[True, False])
        for _, r_ in df_stk_ord.iterrows():
            ws5.append([
                r_.get("producto", ""),
                r_.get("proveedor_dux", ""),
                int(r_["cantidad"]),
                float(r_["costo_unit"]),
                float(r_["valor_total"]),
            ])
        ws5.append(["TOTAL", "", int(df_stock["cantidad"].sum()), "", float(df_stock["valor_total"].sum())])
        last = ws5.max_row
        for c in range(1, 6):
            ws5.cell(row=last, column=c).font = Font(bold=True)
            ws5.cell(row=last, column=c).fill = TOTAL_FILL
    for row_c in ws5.iter_rows(min_row=4, max_col=5):
        for cell in row_c:
            if cell.column in (4, 5):
                cell.number_format = '"$"#,##0'
    for col, w in zip(["A","B","C","D","E"], [55, 30, 12, 15, 15]):
        ws5.column_dimensions[col].width = w
    ws5.freeze_panes = "A4"

    # ======================================================================
    # HOJA 6: COMPRAS HISTÓRICAS
    # ======================================================================
    ws6 = wb.create_sheet("Compras")
    _style_title(ws6, "A1:G1", f"Compras históricas 2024-{hoy.year}")
    ws6.row_dimensions[1].height = 28
    ws6.append([])

    # Resumen por proveedor primero
    resumen_prov = defaultdict(lambda: {"uds": 0, "total": 0, "fechas": []})
    for c in compras:
        p = c.get("prov", "").strip()
        resumen_prov[p]["uds"] += c.get("cant", 0)
        resumen_prov[p]["total"] += c.get("total", 0)
        resumen_prov[p]["fechas"].append(c.get("fecha", ""))

    _style_section(ws6, "A3:G3", "RESUMEN POR PROVEEDOR")
    ws6.append([])
    ws6.append(["Proveedor", "Uds compradas", "Total invertido", "Primera compra", "Última compra", "Cantidad compras", ""])
    _style_header_row(ws6, 5, 7)
    for prov, d in sorted(resumen_prov.items(), key=lambda x: -x[1]["total"]):
        fechas = sorted([f for f in d["fechas"] if f])
        ws6.append([
            prov,
            int(d["uds"]),
            float(d["total"]),
            fechas[0] if fechas else "",
            fechas[-1] if fechas else "",
            len(fechas),
            "",
        ])
    for row_c in ws6.iter_rows(min_row=6, max_col=7):
        for cell in row_c:
            if cell.column == 3:
                cell.number_format = '"$"#,##0'

    # Detalle de compras
    r_start = ws6.max_row + 3
    _style_section(ws6, f"A{r_start}:G{r_start}", "DETALLE COMPLETO DE COMPRAS")
    r_start += 2
    ws6.cell(row=r_start, column=1, value="Fecha")
    ws6.cell(row=r_start, column=2, value="Proveedor")
    ws6.cell(row=r_start, column=3, value="Producto")
    ws6.cell(row=r_start, column=4, value="Cantidad")
    ws6.cell(row=r_start, column=5, value="Costo unitario")
    ws6.cell(row=r_start, column=6, value="Total")
    ws6.cell(row=r_start, column=7, value="Comprobante")
    _style_header_row(ws6, r_start, 7)

    for c in sorted(compras, key=lambda x: (x.get("fecha",""), x.get("prov",""))):
        cant = c.get("cant", 0)
        total = c.get("total", 0)
        cu = total / max(cant, 1) if cant else 0
        ws6.append([
            c.get("fecha", ""),
            c.get("prov", ""),
            c.get("prod", ""),
            int(cant),
            cu,
            total,
            c.get("comp", ""),
        ])
    for row_c in ws6.iter_rows(min_row=r_start+1, max_col=7):
        for cell in row_c:
            if cell.column in (5, 6):
                cell.number_format = '"$"#,##0'
    for col, w in zip(["A","B","C","D","E","F","G"], [12, 30, 50, 12, 14, 14, 18]):
        ws6.column_dimensions[col].width = w

    # ======================================================================
    # HOJA 7: GASTOS 2026
    # ======================================================================
    ws7 = wb.create_sheet(f"Gastos {hoy.year}")
    _style_title(ws7, "A1:F1", f"Gastos {hoy.year}")
    ws7.row_dimensions[1].height = 28
    ws7.append([])

    # Resumen por categoría
    _style_section(ws7, "A3:F3", "RESUMEN POR CATEGORÍA")
    ws7.append([])
    ws7.append(["Categoría", "Total", "Cantidad", "", "", ""])
    _style_header_row(ws7, 5, 3)
    g_anio = [g for g in gastos if g.get("fecha", "").startswith(str(hoy.year))]
    por_cat = defaultdict(lambda: {"total": 0, "cant": 0})
    for g in g_anio:
        por_cat[g.get("categoria", "Otros")]["total"] += g["monto"]
        por_cat[g.get("categoria", "Otros")]["cant"] += 1
    for cat, d in sorted(por_cat.items(), key=lambda x: -x[1]["total"]):
        ws7.append([cat, d["total"], d["cant"], "", "", ""])
    ws7.append(["TOTAL", sum(d["total"] for d in por_cat.values()), sum(d["cant"] for d in por_cat.values()), "", "", ""])
    last = ws7.max_row
    for c in range(1, 4):
        ws7.cell(row=last, column=c).font = Font(bold=True)
        ws7.cell(row=last, column=c).fill = TOTAL_FILL

    # Detalle
    r_start = ws7.max_row + 3
    _style_section(ws7, f"A{r_start}:F{r_start}", "DETALLE DE GASTOS")
    r_start += 2
    ws7.cell(row=r_start, column=1, value="Fecha")
    ws7.cell(row=r_start, column=2, value="Concepto")
    ws7.cell(row=r_start, column=3, value="Categoría")
    ws7.cell(row=r_start, column=4, value="Medio")
    ws7.cell(row=r_start, column=5, value="Monto")
    ws7.cell(row=r_start, column=6, value="Notas")
    _style_header_row(ws7, r_start, 6)
    for g in sorted(g_anio, key=lambda x: x.get("fecha","")):
        ws7.append([
            g.get("fecha", ""),
            g.get("concepto", ""),
            g.get("categoria", ""),
            g.get("medio", ""),
            float(g["monto"]),
            g.get("notas", ""),
        ])
    for row_c in ws7.iter_rows(min_row=6, max_col=6):
        for cell in row_c:
            if cell.column in (2, 5):
                if cell.column == 5 and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0'
    for col, w in zip(["A","B","C","D","E","F"], [12, 55, 22, 18, 14, 40]):
        ws7.column_dimensions[col].width = w

    # ======================================================================
    # Guardar
    # ======================================================================
    return wb


def save_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


if __name__ == "__main__":
    wb = generar_reporte()
    with open("data/reporte_completo.xlsx", "wb") as f:
        f.write(save_to_bytes(wb))
    print("Guardado: data/reporte_completo.xlsx")
